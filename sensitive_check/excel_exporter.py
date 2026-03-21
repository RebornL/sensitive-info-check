"""Excel导出模块

将扫描结果导出为Excel格式的报告。
"""

from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .patterns import Severity, Category
from .scanner import ScanResult, LogSensitiveIssue, FileScanResult


# 样式定义
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL_CRITICAL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
HEADER_FILL_HIGH = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
HEADER_FILL_MEDIUM = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
HEADER_FILL_LOW = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
HEADER_FILL_DEFAULT = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 严重级别排序值
SEVERITY_ORDER = {
    Severity.CRITICAL: 1,
    Severity.HIGH: 2,
    Severity.MEDIUM: 3,
    Severity.LOW: 4,
}


def get_severity_fill(severity: Severity) -> PatternFill:
    """根据严重级别获取填充颜色"""
    fills = {
        Severity.CRITICAL: HEADER_FILL_CRITICAL,
        Severity.HIGH: HEADER_FILL_HIGH,
        Severity.MEDIUM: HEADER_FILL_MEDIUM,
        Severity.LOW: HEADER_FILL_LOW,
    }
    return fills.get(severity, HEADER_FILL_DEFAULT)


def get_severity_label(severity: Severity) -> str:
    """获取严重级别的中文标签"""
    labels = {
        Severity.CRITICAL: "严重(CRITICAL)",
        Severity.HIGH: "高危(HIGH)",
        Severity.MEDIUM: "中级(MEDIUM)",
        Severity.LOW: "低级(LOW)",
    }
    return labels.get(severity, str(severity.label))


def export_to_excel(
    result: ScanResult,
    output_path: str,
    title: Optional[str] = None
) -> str:
    """将扫描结果导出为Excel文件

    Args:
        result: 扫描结果
        output_path: 输出文件路径
        title: 报告标题（可选）

    Returns:
        生成的Excel文件路径
    """
    wb = Workbook()

    # 创建汇总页
    _create_summary_sheet(wb, result, title)

    # 创建详细问题页
    _create_issues_sheet(wb, result)

    # 按严重级别创建分类页
    _create_severity_sheets(wb, result)

    # 删除默认的空Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 保存文件
    wb.save(output_path)
    return output_path


def _create_summary_sheet(wb: Workbook, result: ScanResult, title: Optional[str] = None):
    """创建汇总统计页"""
    ws = wb.active
    ws.title = "汇总报告"

    # 标题
    ws.merge_cells('A1:F1')
    ws['A1'] = title or "敏感信息扫描报告"
    ws['A1'].font = Font(bold=True, size=18, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 扫描时间
    ws.merge_cells('A2:F2')
    ws['A2'] = f"扫描时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=11, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')

    # 空行
    ws.row_dimensions[3].height = 10

    # 扫描统计
    ws['A4'] = "扫描统计"
    ws['A4'].font = TITLE_FONT

    stats_data = [
        ["扫描文件数", result.files_scanned],
        ["发现问题文件数", result.files_with_issues],
        ["问题总数", result.total_issues],
    ]

    for i, (label, value) in enumerate(stats_data, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        for col in ['A', 'B']:
            ws[f'{col}{i}'].border = BORDER

    # 按严重级别统计
    ws['A9'] = "按严重级别统计"
    ws['A9'].font = TITLE_FONT

    headers = ["严重级别", "数量"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_DEFAULT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    severity_data = [
        (Severity.CRITICAL, result.issues_by_severity.get(Severity.CRITICAL, 0)),
        (Severity.HIGH, result.issues_by_severity.get(Severity.HIGH, 0)),
        (Severity.MEDIUM, result.issues_by_severity.get(Severity.MEDIUM, 0)),
        (Severity.LOW, result.issues_by_severity.get(Severity.LOW, 0)),
    ]

    for i, (severity, count) in enumerate(severity_data, start=11):
        ws.cell(row=i, column=1, value=get_severity_label(severity))
        ws.cell(row=i, column=2, value=count)
        ws.cell(row=i, column=1).fill = get_severity_fill(severity)
        ws.cell(row=i, column=1).font = Font(bold=True, color="FFFFFF")
        for col in [1, 2]:
            ws.cell(row=i, column=col).border = BORDER
            ws.cell(row=i, column=col).alignment = Alignment(horizontal='center')

    # 按类别统计
    ws['A16'] = "按敏感信息类别统计"
    ws['A16'].font = TITLE_FONT

    headers = ["类别", "数量"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=17, column=i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_DEFAULT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    row = 18
    for category in Category:
        count = result.issues_by_category.get(category, 0)
        if count > 0:
            ws.cell(row=row, column=1, value=category.value)
            ws.cell(row=row, column=2, value=count)
            for col in [1, 2]:
                ws.cell(row=row, column=col).border = BORDER
            row += 1

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


def _create_issues_sheet(wb: Workbook, result: ScanResult):
    """创建详细问题页"""
    ws = wb.create_sheet("所有问题")

    # 标题
    ws.merge_cells('A1:G1')
    ws['A1'] = "敏感信息检测详情"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25

    # 表头
    headers = ["序号", "严重级别", "类别", "文件路径", "行号", "匹配内容", "修复建议"]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_DEFAULT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 收集所有问题并排序
    all_issues = []
    for file_result in result.file_results:
        for issue in file_result.issues:
            all_issues.append((file_result.file_path, issue))

    # 按严重级别排序
    all_issues.sort(key=lambda x: SEVERITY_ORDER.get(x[1].severity, 99))

    # 填充数据
    row = 3
    for idx, (file_path, issue) in enumerate(all_issues, start=1):
        for match in issue.sensitive_matches:
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=get_severity_label(issue.severity))
            ws.cell(row=row, column=2).fill = get_severity_fill(issue.severity)
            ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=3, value=match.pattern.category.value)
            ws.cell(row=row, column=4, value=file_path)
            ws.cell(row=row, column=5, value=issue.line_number)
            ws.cell(row=row, column=6, value=match.matched_text[:100] if len(match.matched_text) > 100 else match.matched_text)
            ws.cell(row=row, column=7, value=match.pattern.recommendation)

            for col in range(1, 8):
                ws.cell(row=row, column=col).border = BORDER
                ws.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)

            row += 1

    # 设置列宽
    column_widths = [8, 18, 15, 40, 8, 40, 50]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _create_severity_sheets(wb: Workbook, result: ScanResult):
    """按严重级别创建分类页"""
    severity_issues = {
        Severity.CRITICAL: [],
        Severity.HIGH: [],
        Severity.MEDIUM: [],
        Severity.LOW: [],
    }

    # 分类收集问题
    for file_result in result.file_results:
        for issue in file_result.issues:
            if issue.severity in severity_issues:
                severity_issues[issue.severity].append((file_result.file_path, issue))

    # 为每个级别创建Sheet
    sheet_names = {
        Severity.CRITICAL: "严重问题-CRITICAL",
        Severity.HIGH: "高危问题-HIGH",
        Severity.MEDIUM: "中级问题-MEDIUM",
        Severity.LOW: "低级问题-LOW",
    }

    for severity, issues in severity_issues.items():
        if not issues:
            continue

        ws = wb.create_sheet(sheet_names[severity])

        # 标题
        ws.merge_cells('A1:G1')
        ws['A1'] = f"{get_severity_label(severity)}详情"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = get_severity_fill(severity)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25

        # 表头
        headers = ["序号", "类别", "文件路径", "行号", "日志函数", "匹配内容", "修复建议"]
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=i, value=header)
            cell.font = HEADER_FONT
            cell.fill = get_severity_fill(severity)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 填充数据
        row = 3
        for idx, (file_path, issue) in enumerate(issues, start=1):
            for match in issue.sensitive_matches:
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=match.pattern.category.value)
                ws.cell(row=row, column=3, value=file_path)
                ws.cell(row=row, column=4, value=issue.line_number)
                ws.cell(row=row, column=5, value=issue.log_function)
                ws.cell(row=row, column=6, value=match.matched_text[:100] if len(match.matched_text) > 100 else match.matched_text)
                ws.cell(row=row, column=7, value=match.pattern.recommendation)

                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = BORDER
                    ws.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)

                row += 1

        # 设置列宽
        column_widths = [8, 15, 40, 8, 15, 40, 50]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width


def generate_report_filename() -> str:
    """生成默认的报告文件名"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"sensitive_info_report_{timestamp}.xlsx"