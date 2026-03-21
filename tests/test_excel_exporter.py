"""测试Excel导出功能"""

import os
import tempfile
import pytest
from openpyxl import load_workbook

from sensitive_check.patterns import Severity, Category
from sensitive_check.scanner import ScanResult, FileScanResult, LogSensitiveIssue, SensitiveMatch
from sensitive_check.excel_exporter import (
    export_to_excel,
    generate_report_filename,
    get_severity_fill,
    get_severity_label,
)


class TestExcelExporter:
    """测试Excel导出功能"""

    def test_generate_report_filename(self):
        """测试生成报告文件名"""
        filename = generate_report_filename()
        assert filename.startswith("sensitive_info_report_")
        assert filename.endswith(".xlsx")

    def test_get_severity_label(self):
        """测试获取严重级别标签"""
        assert "CRITICAL" in get_severity_label(Severity.CRITICAL)
        assert "HIGH" in get_severity_label(Severity.HIGH)
        assert "MEDIUM" in get_severity_label(Severity.MEDIUM)
        assert "LOW" in get_severity_label(Severity.LOW)

    def test_get_severity_fill(self):
        """测试获取严重级别填充颜色"""
        fill_critical = get_severity_fill(Severity.CRITICAL)
        fill_high = get_severity_fill(Severity.HIGH)
        fill_medium = get_severity_fill(Severity.MEDIUM)
        fill_low = get_severity_fill(Severity.LOW)

        # 不同级别应该有不同的颜色
        assert fill_critical.start_color.rgb != fill_high.start_color.rgb
        assert fill_high.start_color.rgb != fill_medium.start_color.rgb
        assert fill_medium.start_color.rgb != fill_low.start_color.rgb

    def test_export_to_excel_empty_result(self):
        """测试导出空结果"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        try:
            result = ScanResult(
                files_scanned=0,
                files_with_issues=0,
                total_issues=0,
                issues_by_severity={s: 0 for s in Severity},
                issues_by_category={c: 0 for c in Category},
                file_results=[]
            )

            export_to_excel(result, output_path)

            # 验证文件存在
            assert os.path.exists(output_path)

            # 验证文件内容
            wb = load_workbook(output_path)
            assert "汇总报告" in wb.sheetnames

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_to_excel_with_issues(self):
        """测试导出包含问题的结果"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        try:
            # 创建模拟的扫描结果
            from sensitive_check.patterns import SensitivePattern
            import re

            pattern = SensitivePattern(
                name="密码变量",
                category=Category.PASSWORD,
                severity=Severity.CRITICAL,
                pattern=re.compile(r"password\s*=\s*['\"][^'\"]+['\"]"),
                description="检测到密码变量赋值",
                examples=["password = 'secret'"],
                recommendation="不要硬编码密码"
            )

            match = SensitiveMatch(
                line_number=10,
                column_start=1,
                column_end=20,
                matched_text="password = 'secret'",
                pattern=pattern
            )

            issue = LogSensitiveIssue(
                file_path="/test/sample.py",
                line_number=10,
                log_function="print",
                log_content="print(password)",
                sensitive_matches=[match],
                severity=Severity.CRITICAL,
                language="python"
            )

            file_result = FileScanResult(
                file_path="/test/sample.py",
                language="python",
                total_lines=20,
                log_count=1,
                issues=[issue]
            )

            result = ScanResult(
                files_scanned=1,
                files_with_issues=1,
                total_issues=1,
                issues_by_severity={s: 0 for s in Severity},
                issues_by_category={c: 0 for c in Category},
                file_results=[file_result]
            )
            result.issues_by_severity[Severity.CRITICAL] = 1
            result.issues_by_category[Category.PASSWORD] = 1

            export_to_excel(result, output_path)

            # 验证文件存在
            assert os.path.exists(output_path)

            # 验证文件内容
            wb = load_workbook(output_path)

            # 检查Sheet
            assert "汇总报告" in wb.sheetnames
            assert "所有问题" in wb.sheetnames
            assert "严重问题-CRITICAL" in wb.sheetnames

            # 检查汇总页内容
            ws = wb["汇总报告"]
            assert ws['A1'].value == "敏感信息扫描报告"

            # 检查问题页内容
            ws_issues = wb["所有问题"]
            assert ws_issues['A2'].value == "序号"
            assert ws_issues['B2'].value == "严重级别"
            assert ws_issues['C2'].value == "类别"

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_to_excel_multiple_severities(self):
        """测试导出多种严重级别的结果"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        try:
            # 创建不同严重级别的问题
            from sensitive_check.patterns import SensitivePattern
            import re

            issues = []
            severities = [Severity.CRITICAL, Severity.HIGH, Severity.MEDIUM, Severity.LOW]
            categories = [Category.PASSWORD, Category.ID_CARD, Category.PHONE, Category.IP_ADDRESS]

            for i, (sev, cat) in enumerate(zip(severities, categories)):
                pattern = SensitivePattern(
                    name=f"测试模式{i}",
                    category=cat,
                    severity=sev,
                    pattern=re.compile(r"test"),
                    description=f"测试描述{i}",
                    examples=["test"],
                    recommendation=f"测试建议{i}"
                )

                match = SensitiveMatch(
                    line_number=i + 1,
                    column_start=1,
                    column_end=10,
                    matched_text=f"test_match_{i}",
                    pattern=pattern
                )

                issue = LogSensitiveIssue(
                    file_path=f"/test/file{i}.py",
                    line_number=i + 1,
                    log_function="print",
                    log_content="print(test)",
                    sensitive_matches=[match],
                    severity=sev,
                    language="python"
                )
                issues.append(issue)

            file_result = FileScanResult(
                file_path="/test/",
                language="python",
                total_lines=100,
                log_count=4,
                issues=issues
            )

            result = ScanResult(
                files_scanned=1,
                files_with_issues=1,
                total_issues=4,
                issues_by_severity={s: 1 for s in Severity},
                issues_by_category={c: 1 for c in Category},
                file_results=[file_result]
            )

            export_to_excel(result, output_path)

            # 验证文件存在
            assert os.path.exists(output_path)

            # 验证Sheet
            wb = load_workbook(output_path)
            assert "严重问题-CRITICAL" in wb.sheetnames
            assert "高危问题-HIGH" in wb.sheetnames
            assert "中级问题-MEDIUM" in wb.sheetnames
            assert "低级问题-LOW" in wb.sheetnames

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_with_custom_title(self):
        """测试使用自定义标题导出"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name

        try:
            result = ScanResult(
                files_scanned=0,
                files_with_issues=0,
                total_issues=0,
                issues_by_severity={s: 0 for s in Severity},
                issues_by_category={c: 0 for c in Category},
                file_results=[]
            )

            export_to_excel(result, output_path, title="自定义报告标题")

            wb = load_workbook(output_path)
            ws = wb["汇总报告"]
            assert ws['A1'].value == "自定义报告标题"

        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)